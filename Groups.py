'''
Created on March 27, 2024
@author: Matthew Pryke
'''

#================================== Imports ===================================



import pandas as pd
from openpyxl.workbook import Workbook
from pandas.core.computation import align
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
import xlsxwriter
from xlsxwriter.workbook import Workbook as xlsxWorkbook
import os
import time
import datetime



class GroupCreation:
    def __init__(self) -> None:
        self.competition_name = ""
        self.player_numbers = False
        self.seeding_directory = "output files/seedings.xlsx"
        self.players = []

        # create_groups stuff
        self.group_number = 0
        self.forward = True
        self.end = 1

        # Writing to excel file stuff
        self.row = 1



    def get_competition_name(self) -> str:
        '''
        Gets the competition name
        :return: str
        '''
        return self.competition_name



    def __date(self) -> str:
        '''
        Gets the date from the user and converts it to the full date
        :return: str
        '''

        # date = input(f"What date is the {event} taking place? (DD/MM/YYYY): ")
        date = "27/03/2024"
        day_number = int(date[0:2])
        month = int(date[3:5])
        year = int(date[-4:])
        date = datetime.datetime(int(date[-4:]), int(date[3:5]), int(date[0:2]))

        # day of the week
        day = date.weekday()
        match day:
            case 0:
                day = "Monday"
            case 1:
                day = "Tuesday"
            case 2:
                day = "Wednesday"
            case 3:
                day = "Thursday"
            case 4:
                day = "Friday"
            case 5:
                day = "Saturday"
            case 6:
                day = "Sunday"

        # suffix for the day of the month
        if day_number == 1 or day_number == 21 or day_number == 31:
            number_suffix = "st"
        elif day_number == 2 or day_number == 22:
            number_suffix = "nd"
        elif day_number == 3 or day_number == 23:
            number_suffix = "rd"
        else:
            number_suffix = "th"

        # month
        match month:
            case 1:
                month = "January"
            case 2:
                month = "Febuary"
            case 3:
                month = "March"
            case 4:
                month = "April"
            case 5:
                month = "May"
            case 6:
                month = "June"
            case 7:
                month = "July"
            case 8:
                month = "August"
            case 9:
                month = "September"
            case 10:
                month = "October"
            case 11:
                month = "November"
            case 12:
                month = "December"

        # put it all together
        return f"{day} {day_number}{number_suffix} {month} {year}"



    def __get_player_list(self, player_excel: Workbook, event: str) -> None:
        ws = player_excel[event]

        counter = 2
        while True:
            person = []
            if ws["A" + str(counter)].value is not None:
                person.append(ws["A" + str(counter)].value)
                person.append(ws["B" + str(counter)].value)
                person.append(ws["C" + str(counter)].value)
                person.append(ws["D" + str(counter)].value)
                if self.player_numbers:
                    person.append(ws["E" + str(counter)].value)
                self.players.append(person)
                counter += 1
            else:
                break



    def __number_of_groups(self, prefered_group_size: int) -> int | tuple[int, int]:
        '''
        Calculates the number of groups needed for the event
        :param prefered_group_size: integer representing the number prefered group size
        :return: tuple [number of groups, max group size]
        '''

        number_of_players = len(self.players)
        number_of_players = 37

        # Even number of people in each group
        if number_of_players % prefered_group_size == 0:
            return number_of_players // prefered_group_size

        # Odd number of people in each group
        above_prefered_max = input("""It is not possible to have the same number of people in each group. Would you like to go above the prefered group size?: (y/n)""")
        if above_prefered_max.lower() == "y" or above_prefered_max.lower() == "yes":
            return number_of_players // prefered_group_size, prefered_group_size + 1
        else:
            return number_of_players // prefered_group_size + 1, prefered_group_size



    def __create_groups(self, number_of_groups: int, max_group_size: int) -> int:
        '''
        Creates the groups for the competition vai the sanke method
        :param number_of_groups: the number of groups
        :param max_group_size: the max size of a group
        :returns: The largest group size
        '''

        # Initializies the groups
        groups = []
        group_length = [0] * number_of_groups
        for i in range(number_of_groups):
            groups.append([])

        previous_group_number = 0
        clash_moved_to = 1234567890


        for player in self.players:
            # need to get it to skip the group where it placed the clashed person
            # check the previous group
            if clash_moved_to == self.group_number:
                clash_moved_to = 123456789
                self.__change_group(number_of_groups)

            # if the current group is the same as the previous group then skip it

            # Check if there is a county clash
            clash = False
            for player_in_group in groups[self.group_number]:
                    if player_in_group[2] == player[2]:
                        clash = True
                        # print(f"There are {group_length[self.group_number]} players")
                        # print(f"clash in group {self.group_number}")
                        break


            # Trys to change the group it is in when there is a clash
            if clash:
                original_group = self.group_number
                original_forward = self.forward
                original_end = self.end
                tries = 0
                while clash:
                    # move to the next group
                    self.__change_group(number_of_groups)
                    if self.group_number == original_group:
                        continue

                    # Makes sure that the next group isnt full
                    if len(groups[self.group_number]) != max_group_size:

                        # Checks for clash in new group
                        clash_new_group = False
                        for player_in_group in groups[self.group_number]:
                            if player_in_group[2] == player[2]:
                                clash_new_group = True
                                # print(f"clash in group {self.group_number}")

                        if not clash_new_group:
                            groups[self.group_number].append(player)
                            group_length[self.group_number] += 1
                            break
                        else:
                            tries += 1
                            if tries == number_of_groups - 1:
                                groups[original_group].append(player)
                                group_length[original_group] += 1
                                break
                    else:
                        tries += 1

                clash_moved_to = self.group_number
                self.group_number = original_group
                self.forward = original_forward
                self.end = original_end


            else:
                # Makes sure the group isnt full
                while (group_length[self.group_number] == max_group_size):
                    previous_group_number = self.group_number
                    self.__change_group(number_of_groups)

                # Adds the player to the group
                groups[self.group_number].append(player)
                group_length[self.group_number] += 1
                previous_group_number = self.group_number
                self.__change_group(number_of_groups)


        # for group in groups:
        #     print(group)
        #     print(len(group))

        return max(group_length)



    def __change_group(self, number_of_groups: int) -> int:
        '''
        Changes the group number in accordance with the snake method
        :param group_number: The current group number
        :param forward: The current direction the snake is going
        :param end: Whether the snake has done the end group twice or not
        :param number_of_groups: how many groups are in the competition
        :param clash: whether it is in the clash state or not
        :return: The next group number
        '''
        # change the group number
        if self.forward:
            if self.group_number == (number_of_groups - 1) and self.end == 2:
                self.forward = False
                self.group_number -= 1
                self.end = 1
            elif self.group_number == (number_of_groups - 1) and self.end == 1:
                self.end = 2
            else:
                self.group_number += 1
        else:
            if self.group_number == 0 and self.end == 2:
                self.forward = True
                self.group_number += 1
                self.end = 1
            elif self.group_number == 0 and self.end == 1:
                self.end = 2
            else:
                self.group_number -= 1



    def __write_file(self, group_file: xlsxWorkbook, sheet: xlsxwriter.workbook.Worksheet, number_of_groups: int, largest_group_size: int) -> None:
        '''
        Writes the groups to the file
        :param groups_file:
        :return:
        '''

        # Creates the headings
        heading_format = group_file.add_format({"bold": True, "border": 1, "top": 2, "align": "center"})
        sheet.write_string(self.row, 0, "Date", cell_format=heading_format)
        sheet.write_string(self.row, 1, "Event", cell_format=heading_format)
        sheet.write_string(self.row, 2, "Group", cell_format=heading_format)

        col = 3
        for i in range(largest_group_size):
            if self.player_numbers:
                sheet.write_string(self.row, col, f"No{chr(i + 65)}", cell_format=heading_format)
                col += 1
            sheet.write_string(self.row, col, f"Cod{chr(i + 65)}", cell_format=heading_format)
            col += 1
            sheet.write_string(self.row, col, f"Player{chr(i + 65)}", cell_format=heading_format)
            col += 1




    def main(self) -> None:
        '''
        Main function that creates the groups by calling the other functions in the class.
        :return: None
        '''
        # Inputs
        while True:
            self.competition_name = input("Enter competition name: ")
            if self.competition_name != "":
                break

        # while True:
        #     player_numbers = input("Are player numbers in use at this event? (y/n): ")
        #     if player_numbers.lower() == "y" or player_numbers.lower() == "yes":
        #         self.player_numbers = True
        #         break
        #     elif player_numbers.lower() == "n" or player_numbers.lower() == "no":
        #         break
        #     else:
        #         print("Sorry, I didn't understand")

        # Loads the seeding File and grabs its Sheets
        try:
            player_excel = load_workbook(self.seeding_directory)
        except FileNotFoundError:
            print(f"{bcolors.WARNING}Error code 3: Seeding list not Found{bcolors.ENDC}")
            time.sleep(3)
            exit(3)
        events = load_workbook(self.seeding_directory).sheetnames

        groups = xlsxwriter.Workbook("output files/Groups.xlsx")
        sheet = groups.add_worksheet("Groups")
        sheet.merge_range("A1:G1", f"{self.competition_name}", cell_format=groups.add_format({'bold': True}))
        sheet.write_string(1, 0, "test")



        # Looks at all the sheets in the Excel file
        for idx, event in enumerate(events):
            # Date
            date = self.__date()

            # Player info
            self.__get_player_list(player_excel, event)

            # Group size
            print(f"There are {len(self.players)} in this event")
            prefered_group_size = int(input("Enter prefered group size: "))

            # Calculate number of groups
            number_of_groups, max_group_size = self.__number_of_groups(prefered_group_size)

            # Create Groups
            largest_group_size = self.__create_groups(number_of_groups, max_group_size)
            self.__write_file(groups, sheet, number_of_groups, largest_group_size)



        groups.close()
        player_excel.close()


# player numbers
# size of groups wanted
# event

# player[0] = licence number
# player[1] = name
# player[2] = county
# player[3] = points

if __name__ == "__main__":
    GroupCreation().main()