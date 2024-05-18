'''
Created on March 27, 2024
@author: Matthew Pryke
'''

#================================== Imports ===================================


from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import xlsxwriter
from xlsxwriter.workbook import Workbook as xlsxWorkbook
import os
import time
import datetime
from docxtpl import DocxTemplate
import shutil
from utils.text_colour import bcolors
import main


class GroupCreation:
    def __init__(self) -> None:
        self.competition_name = ""
        self.player_numbers = False
        self.seeding_directory = "output files/seedings.xlsx"
        self.date = ""
        self.players = []

        # create_groups stuff
        self.groups = []
        self.group_number = 0
        self.forward = True
        self.end = 1

        # Writing to excel file stuff
        self.row = 1

        # Group sheet stuff
        self.short_date = ""



    def get_competition_name(self) -> str:
        '''
        Gets the competition name
        :return: str
        '''
        return self.competition_name



    def __date(self, event: str) -> str:
        '''
        Gets the date from the user and converts it to the full date
        :return: str
        '''

        while True:
            date = input(f"What date is the {event} taking place? (DD/MM/YYYY): ")
            # date = "27/03/2024"
            self.short_date = date
            try:
                day_number = int(date[0:2])
                month = int(date[3:5])
                year = int(date[-4:])
                date = datetime.datetime(int(date[-4:]), int(date[3:5]), int(date[0:2]))
            except Exception:
                print("Error 4: Incorrect format entered. Please try again.")
            else:
                break

        # day of the week
        day = date.weekday()
        match day:
            case 0:
                day = "Monday"
            case 1:
                day = "Tuesday"
            case 2:
                day = "Wednesday"
            case 3:
                day = "Thursday"
            case 4:
                day = "Friday"
            case 5:
                day = "Saturday"
            case 6:
                day = "Sunday"

        # suffix for the day of the month
        if day_number == 1 or day_number == 21 or day_number == 31:
            number_suffix = "st"
        elif day_number == 2 or day_number == 22:
            number_suffix = "nd"
        elif day_number == 3 or day_number == 23:
            number_suffix = "rd"
        else:
            number_suffix = "th"

        # month
        match month:
            case 1:
                month = "January"
            case 2:
                month = "February"
            case 3:
                month = "March"
            case 4:
                month = "April"
            case 5:
                month = "May"
            case 6:
                month = "June"
            case 7:
                month = "July"
            case 8:
                month = "August"
            case 9:
                month = "September"
            case 10:
                month = "October"
            case 11:
                month = "November"
            case 12:
                month = "December"

        # put it all together
        return f"{day} {day_number}{number_suffix} {month} {year}"



    def __get_player_list(self, player_excel: Workbook, event: str) -> None:
        ws = player_excel[event]
        self.players = []

        counter = 2
        while True:
            person = []
            if ws["A" + str(counter)].value is not None:
                person.append(ws["A" + str(counter)].value)
                person.append(ws["B" + str(counter)].value)
                person.append(ws["C" + str(counter)].value)
                person.append(ws["D" + str(counter)].value)
                if self.player_numbers:
                    person.append(ws["E" + str(counter)].value)
                self.players.append(person)
                counter += 1
            else:
                break



    def __number_of_groups(self, prefered_group_size: int) -> int | tuple[int, int]:
        '''
        Calculates the number of groups needed for the event
        :param prefered_group_size: integer representing the number prefered group size
        :return: tuple [number of groups, max group size]
        '''

        number_of_players = len(self.players)
        #number_of_players = 37

        # Even number of people in each group
        if number_of_players % prefered_group_size == 0:
            return number_of_players // prefered_group_size, prefered_group_size

        if number_of_players // prefered_group_size == 1:
            return (number_of_players // prefered_group_size) + 1, prefered_group_size

        # Odd number of people in each group
        above_prefered_max = input("""It is not possible to have the same number of people in each group. Would you like to go above the prefered group size?: (y/n)""")
        if above_prefered_max.lower() == "y" or above_prefered_max.lower() == "yes":
            return number_of_players // prefered_group_size, prefered_group_size + 1
        else:
            return (number_of_players // prefered_group_size) + 1, prefered_group_size



    def __create_groups(self, number_of_groups: int, max_group_size: int) -> int:
        '''
        Creates the groups for the competition vai the sanke method
        :param number_of_groups: the number of groups
        :param max_group_size: the max size of a group
        :returns: The largest group size
        '''

        # Initializies the groups
        #self.group_number = 0
        self.groups = []
        group_length = [0] * number_of_groups
        for i in range(number_of_groups):
            self.groups.append([])

        previous_group_number = 0
        clash_moved_to = 1234567890


        for player in self.players:
            # need to get it to skip the group where it placed the clashed person
            # check the previous group
            if clash_moved_to == self.group_number:
                clash_moved_to = 123456789
                self.__change_group(number_of_groups)

            # if the current group is the same as the previous group then skip it

            # Check if there is a county clash
            clash = False
            for player_in_group in self.groups[self.group_number]:
                    if player_in_group[2] == player[2]:
                        clash = True
                        # print(f"There are {group_length[self.group_number]} players")
                        # print(f"clash in group {self.group_number}")
                        break


            # Trys to change the group it is in when there is a clash
            if clash:
                original_group = self.group_number
                original_forward = self.forward
                original_end = self.end
                tries = 0
                while clash:
                    # move to the next group
                    self.__change_group(number_of_groups)
                    if self.group_number == original_group:
                        continue

                    # Makes sure that the next group isnt full
                    if len(self.groups[self.group_number]) != max_group_size:

                        # Checks for clash in new group
                        clash_new_group = False
                        for player_in_group in self.groups[self.group_number]:
                            if player_in_group[2] == player[2]:
                                clash_new_group = True
                                # print(f"clash in group {self.group_number}")

                        if not clash_new_group:
                            self.groups[self.group_number].append(player)
                            group_length[self.group_number] += 1
                            break
                        else:
                            tries += 1
                            if tries == number_of_groups - 1:
                                self.groups[original_group].append(player)
                                self.group_length[original_group] += 1
                                break
                    else:
                        tries += 1

                clash_moved_to = self.group_number
                self.group_number = original_group
                self.forward = original_forward
                self.end = original_end


            else:
                # Makes sure the group isnt full
                while (group_length[self.group_number] == max_group_size):
                    previous_group_number = self.group_number
                    self.__change_group(number_of_groups)

                # Adds the player to the group
                self.groups[self.group_number].append(player)
                group_length[self.group_number] += 1
                previous_group_number = self.group_number
                self.__change_group(number_of_groups)


        # for group in groups:
        #     print(group)
        #     print(len(group))

        return max(group_length)



    def __change_group(self, number_of_groups: int) -> int:
        '''
        Changes the group number in accordance with the snake method
        :param group_number: The current group number
        :param forward: The current direction the snake is going
        :param end: Whether the snake has done the end group twice or not
        :param number_of_groups: how many groups are in the competition
        :param clash: whether it is in the clash state or not
        :return: The next group number
        '''
        # change the group number
        if self.forward:
            if self.group_number == (number_of_groups - 1) and self.end == 2:
                self.forward = False
                self.group_number -= 1
                self.end = 1
            elif self.group_number == (number_of_groups - 1) and self.end == 1:
                self.end = 2
            else:
                self.group_number += 1
        else:
            if self.group_number == 0 and self.end == 2:
                self.forward = True
                self.group_number += 1
                self.end = 1
            elif self.group_number == 0 and self.end == 1:
                self.end = 2
            else:
                self.group_number -= 1



    def __write_excel(self, group_file: xlsxWorkbook, sheet: xlsxwriter.workbook.Worksheet, number_of_groups: int, largest_group_size: int, event: str) -> None:
        '''
        Writes the groups to the file
        :param groups_file:
        :param sheet:
        :param number_of_groups:
        :param largest_group_size:
        :param event:
        :return:
        '''
        # All the formats needed
        heading_format = group_file.add_format({"bold": True, "border": 1, "top": 2, "align": "center", "valign": "vcenter"})
        date_event_format = group_file.add_format({"bold": True, "border": 1, "valign": "vcenter"})
        number_format = group_file.add_format({"bold": True, "border": 1, "align": "center", "valign": "vcenter"})
        licence_number_format = group_file.add_format({"border": 1, "align": "center", "valign": "vcenter"})
        player_name_format = group_file.add_format({"border": 1, "valign": "vcenter"})

        date_event_last_group_format = group_file.add_format({"bold": True, "border": 1, "bottom": 2, "valign": "vcenter"})
        number_last_group_format = group_file.add_format({"bold": True, "border": 1, "bottom": 2, "align": "center", "valign": "vcenter"})
        licence_number_last_group_format = group_file.add_format({"border": 1, "bottom": 2, "align": "center", "valign": "vcenter"})
        player_name_last_group_format = group_file.add_format({"border": 1, "bottom": 2, "valign": "vcenter"})
        player_name_last_group_name_format = group_file.add_format({"border": 1, "bottom": 2, "right": 2, "valign": "vcenter"})
        player_name_last_name_format = group_file.add_format({"border": 1, "right": 2, "valign": "vcenter"})

        heading_end_format = group_file.add_format({"bold": True, "border": 1, "top": 2, "right": 2, "align": "center", "valign": "vcenter"})

        # Creates the headings
        sheet.set_row(self.row, 19.5)
        sheet.write_string(self.row, 0, "Date", cell_format=heading_format)
        sheet.write_string(self.row, 1, "Event", cell_format=heading_format)
        sheet.write_string(self.row, 2, "Group", cell_format=heading_format)

        col = 3
        for i in range(largest_group_size):
            if self.player_numbers:
                sheet.write_string(self.row, col, f"No{chr(i + 65)}", cell_format=heading_format)
                col += 1
            sheet.write_string(self.row, col, f"Cod{chr(i + 65)}", cell_format=heading_format)
            col += 1
            sheet.write_string(self.row, col, f"Player{chr(i + 65)}", cell_format=heading_format)
            col += 1
            sheet.write_string(self.row, col, f"c{chr(i + 65)}", cell_format=heading_format if i + 1 != largest_group_size else heading_end_format)
            col += 1

        self.row += 1

        # writes the date
        sheet.write_string(self.row, 0, self.date, cell_format=date_event_format)

        # writes the Groups
        for i in range(number_of_groups):
            if i  != 0:
                sheet.write_string(self.row, 0, "", cell_format=player_name_format if i + 1 != number_of_groups else player_name_last_group_format)
            sheet.write_string(self.row, 1, event, cell_format=date_event_format if i + 1 != number_of_groups else date_event_last_group_format)
            sheet.write_number(self.row, 2, i + 1, cell_format=number_format if i + 1 != number_of_groups else number_last_group_format)
            col = 3

            player_count = 0
            for player in self.groups[i]:
                # writes the player number
                if self.player_numbers:
                    sheet.write_number(self.row, col, 1, cell_format=number_format if i + 1 != number_of_groups else number_last_group_format)
                    col += 1

                # writes the licence number
                sheet.write_number(self.row, col, player[0], cell_format=licence_number_format if i + 1 != number_of_groups else licence_number_last_group_format)
                col += 1

                # writes player name
                sheet.write_string(self.row, col, player[1], cell_format=player_name_format if i + 1 != number_of_groups else player_name_last_group_format)
                col += 1
                player_count += 1

                # writes the county
                if player_count == largest_group_size:
                    format = player_name_last_name_format if i + 1 != number_of_groups else player_name_last_group_name_format
                else:
                    format = player_name_format if i + 1 != number_of_groups else player_name_last_group_format
                sheet.write_string(self.row, col, player[2], cell_format=format)
                col += 1

            # fills in the blanks if needed
            while player_count != largest_group_size:
                if self.player_numbers:
                    sheet.write_string(self.row, col, "", cell_format=number_format if i + 1 != number_of_groups else number_last_group_format)
                    col += 1

                sheet.write_string(self.row, col, "", cell_format=licence_number_format if i + 1 != number_of_groups else licence_number_last_group_format)
                col += 1

                sheet.write_string(self.row, col, "", cell_format=player_name_format if i + 1 != number_of_groups else player_name_last_group_format)
                col += 1

                if player_count == largest_group_size - 1:
                    format = player_name_last_name_format if i + 1 != number_of_groups else player_name_last_group_name_format
                else:
                    format = player_name_format if i + 1 != number_of_groups else player_name_last_group_format
                sheet.write_string(self.row, col, "", cell_format=format)
                col += 1

                player_count += 1

            self.row += 1
        self.row += 1



    def __group_sheets(self, event: str) -> None:
        '''
        Creates the groups sheets
        :param event:
        :return:
        '''

        try:
            os.listdir("output files/group sheets")
        except FileNotFoundError:
            os.mkdir("output files/group sheets")

        # creates the event folder if not exists
        try:
            shutil.rmtree(f"output files/group sheets/{event}")
        except FileNotFoundError:
            os.mkdir(f"output files/group sheets/{event}")
        else:
            os.mkdir(f"output files/group sheets/{event}")

        # creates the sheets for all the groups in the event
        for i, group in enumerate(self.groups):
            group_size = len(self.groups[i])

            # if self.player_numbers:
            #     pass
            # else:
            #     shutil.copy(f"assets/group_sheets/{group_size}.docx", f"output files/group sheets/{event}/{i + 1}.docx")


            # creates the keys for the dictionay
            keys = ["competition_name", "event_name", "event_date", "group_number"]
            for x in range(group_size):
                if self.player_numbers:
                    keys.append(f"player_number{chr(x + 65)}")
                keys.append(f"cod{chr(x + 65)}")
                keys.append(f"Player{chr(x + 65)}")
                keys.append(f"c{chr(x + 65)}")

            # puts the data into a single list
            data = [self.competition_name, event, self.short_date, str(i + 1)]
            for player in group:
                if self.player_numbers:
                    pass
                data.append(player[0])
                data.append(player[1])
                data.append(player[2])

            # creates the dictionary
            dictionary = {}
            for x, key in enumerate(keys):
                dictionary[key] = data[x]

            # edits the file accordingly
            document = DocxTemplate(f"resources/group_sheets/{group_size}.docx")

            document.render(dictionary)
            document.save(f"output files/group sheets/{event}/{i + 1}.docx")




    def main(self) -> None:
        '''
        Main function that creates the groups by calling the other functions in the class.
        :return: None
        '''
        # Inputs
        while True:
            self.competition_name = input("Enter competition name: ")
            if self.competition_name != "":
                break

        # while True:
        #     player_numbers = input("Are player numbers in use at this event? (y/n): ")
        #     if player_numbers.lower() == "y" or player_numbers.lower() == "yes":
        #         self.player_numbers = True
        #         break
        #     elif player_numbers.lower() == "n" or player_numbers.lower() == "no":
        #         break
        #     else:
        #         print("Sorry, I didn't understand")

        # Loads the seeding File and grabs its Sheets
        try:
            player_excel = load_workbook(self.seeding_directory)
        except FileNotFoundError:
            print(f"{bcolors.WARNING}Error code 3: Seeding list not Found{bcolors.ENDC}")
            time.sleep(3)
            exit(3)
        events = load_workbook(self.seeding_directory).sheetnames

        groups = xlsxwriter.Workbook("output files/Groups.xlsx")
        sheet = groups.add_worksheet("Groups")
        sheet.set_landscape()
        sheet.merge_range("A1:G1", f"{self.competition_name}", cell_format=groups.add_format({'bold': True, "valign": "vcenter"}))
        sheet.set_row(0, 19.5)


        # Looks at all the sheets in the Excel file
        for idx, event in enumerate(events):
            # Date
            self.date = self.__date(event)

            # Player info
            self.__get_player_list(player_excel, event)

            # Group size
            print(f"There are {len(self.players)} in this event")
            prefered_group_size = int(input("Enter preferred group size: "))

            # Calculate number of groups
            number_of_groups, max_group_size = self.__number_of_groups(prefered_group_size)

            # Create Groups
            largest_group_size = self.__create_groups(number_of_groups, max_group_size)

            # Writes groups to excel
            self.__write_excel(groups, sheet, number_of_groups, largest_group_size, event)
            self.__group_sheets(event)


        sheet.autofit()
        groups.close()
        player_excel.close()

        main.Main().menu()


# player numbers
# size of groups wanted
# event

# player[0] = licence number
# player[1] = name
# player[2] = county
# player[3] = points

if __name__ == "__main__":
    GroupCreation().main()