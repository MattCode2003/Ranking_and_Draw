import os
import shutil
import pandas as pd
import time
import xlsxwriter
from docxtpl import DocxTemplate
from openpyxl.reader.excel import load_workbook
from sqlalchemy import create_engine, text
import datetime


class Controller:

#===================================================== Initialiser =====================================================



    def __init__(self, model, view) -> None:
        self.model = model
        self.view = view


        # Turns Ranking List Excel File into Database stored in RAM
        self.engine = create_engine('sqlite:///', echo=False)
        df = pd.read_excel(self.__get_model().get_alpha_list())
        df.to_sql('ranking_list', con=self.engine, if_exists='replace', index=False)

        # Turns the County codes excel file into database stored in RAM
        try:
            df = pd.read_excel("resources/County Codes.xlsx")
        except FileNotFoundError:
            print("Error code 2: County_Codes file not found")
            time.sleep(3)
            exit(2)
        finally:
            df.to_sql('county_codes', con=self.engine, if_exists='replace', index=False)


        # Starts the main menu
        self.__menu_choice(1)



#===================================================== Menu Choice =====================================================



    def __menu_choice(self, menu: int) -> None:
        match menu:

            # Main Menu
            case 1:
                user_choice = self.__get_view().main_menu()
                match user_choice:

                    # Ranking
                    case 1:
                        self.__menu_choice(2)

                    # Groups
                    case 2:
                        self.__groups()

                    # Exit
                    case 3:
                        exit(0)

            # Ranking Menu
            case 2:
                user_choice = (self.__get_view().ranking_menu())
                match user_choice:

                    # Single player Ranking
                    case 1:
                        self.__single_player_ranking()

                    # Entire Tournament rankings
                    case 2:
                        self.__entire_tournament_ranking()

                    # Main Menu
                    case 3:
                        self.__menu_choice(1)



#================================================ Single Player Ranking ================================================



    def __single_player_ranking(self) -> None:
        player_id = self.__get_view().single_ranking_id()

        with self.__get_engine().connect() as conn:
            query = f"SELECT Sub_Category, Points FROM ranking_list WHERE Membership_no = '{player_id}';"
            result = conn.execute(text(query)).fetchall()

        self.__get_view().output_single_player_ranking(result)
        self.__menu_choice(2)



#============================================== Entire Tournament Ranking ==============================================



    def __get_player_list(self, player_excel, event: str) -> None:

        ws = player_excel[event]


        # Skip age group if it is not taking place
        if ws["A2"].value is None:
            return

        # In every Sheet it takes all the names and licence numbers in each sheet and puts it in a list
        players = self.__get_model().get_players()
        counter = 2
        while True:
            person = []
            if ws["A" + str(counter)].value is not None:
                person.append(ws["A" + str(counter)].value)
                person.append(ws["B" + str(counter)].value)
                person.append(ws["C" + str(counter)].value)
                players.append(person)
                counter += 1
            else:
                break
        self.__get_model().set_players(players)



    def __event(self, event: str) -> str:
        # Works out what Age Category the Players are in
        players = self.__get_model().get_players()
        current_event = None
        match event:
            case "u11b":
                current_event = "Under 11 Boys"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u11g":
                current_event = "Under 11 Girls"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u13b":
                current_event = "Under 13 Boys"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u13g":
                current_event = "Under 13 Girls"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "cadet boys":
                current_event = "Under 15 Boys"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "cadet girls":
                current_event = "Under 15 Girls"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "jnr boys":
                current_event = "Under 19 Men"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "jnr girls":
                current_event = "Under 19 Women"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u21m":
                current_event = "Under 21 Men"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u21w":
                current_event = "Under 21 Women"
                for player in players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "mens singles":
                current_event = "Mens Singles"
                for player in players:
                    player.append(self.__get_points(player[0], "Men", "Senior"))

            case "womens singles":
                current_event = "Womens Singles"
                for player in players:
                    player.append(self.__get_points(player[0], "Women", "Senior"))

            case "mens vets":
                current_event = "Mens Vets Singles"
                for player in players:
                    player.append(self.__get_points(player[0], "Men", "Veteran"))

            case "womens vets":
                current_event = "Womens Vets Singles"
                for player in players:
                    player.append(self.__get_points(player[0], "Women", "Veteran"))

        self.__get_model().set_players(players)
        return current_event



    def __get_points(self, licence_number: str, event: str, senior_category: str) -> int:
        with self.__get_engine().connect() as conn:
            if senior_category != "":
                query = f"SELECT Points FROM ranking_list WHERE Membership_no = '{licence_number}' AND Sub_Category = '{event}' AND Category = '{senior_category}';"
            else:
                query = f"SELECT Points FROM ranking_list WHERE Membership_no = '{licence_number}' AND Sub_Category = '{event}';"
            points = conn.execute(text(query)).fetchall()

            if points == []:
                just_points = "0"
            else:
                just_points = ""
                rb = list(str(points))
                for character in rb:
                    if character.isnumeric():
                        just_points += character

        return int(just_points)



    def __county_codes(self) -> None:
        # Converts the counties into the county codes
        players = self.__get_model().get_players()
        for player in players:
            with self.__get_engine().connect() as conn:
                query = f"SELECT Code FROM county_codes WHERE County = '{player[2]}';"
                county_code = conn.execute(text(query)).fetchall()

            # Removes the extra crap that is in the county code string
            just_county_code = ""
            scounty = list(str(county_code))
            for char in scounty:
                if char.isalpha():
                    just_county_code += char

            player[2] = just_county_code

        self.__get_model().set_players(players)



    def __write_to_file(self, seedings, current_event: str) -> None:
        sheet = seedings.add_worksheet(current_event)
        sheet.write_string(0, 0, "Licence Number")
        sheet.write_string(0, 1, "Name")
        sheet.write_string(0, 2, "County")
        sheet.write_string(0, 3, "Points")

        players = self.__get_model().get_players()

        for i, player in enumerate(players):
            for x, data in enumerate(player):
                if x == 0 or x == 3:
                    sheet.write_number(i + 1, x, data)
                else:
                    sheet.write_string(i + 1, x, data)

        # Auto changes the size of the cells to fit perfectly
        sheet.autofit()



    def __entire_tournament_ranking(self) -> None:
        # Loads the Player List File and grabs its Sheets
        player_excel = load_workbook(self.__get_model().get_player_list())
        events = load_workbook(self.__get_model().get_player_list()).sheetnames

        # Creates the excel file for the seedings
        seedings = xlsxwriter.Workbook("output files/seedings.xlsx")

        # Looks at all the sheets in the Excel file
        for idx, event in enumerate(events):

            # Skip age group if it is not taking place
            ws = player_excel[event]
            if ws["A2"].value is None:
                continue

            self.__get_model().set_players([])
            self.__get_player_list(player_excel, event)
            current_event = self.__event(event)

            # Sorts the players in order of most points
            players = self.__get_model().get_players()
            players.sort(key=lambda x: x[3], reverse=True)
            self.__get_model().set_players(players)
            self.__county_codes()
            self.__write_to_file(seedings, current_event)

        # if there is no output folder then one is created
        path = "output files"
        if not os.path.exists(path):
            os.makedirs(path)
        seedings.close()
        player_excel.close()

        self.__get_model().set_players([])

        self.__menu_choice(2)



#=================================================== Group Creation ====================================================



    def __date(self, event: str) -> str:
        '''
        Gets the date from the user and converts it to the full date
        :return: str
        '''

        while True:
            date = self.__get_view().date(event)
            # date = "27/03/2024"
            self.__get_model().set_short_date(date)
            try:
                day_number = int(date[0:2])
                month = int(date[3:5])
                year = int(date[-4:])
                date = datetime.datetime(int(date[-4:]), int(date[3:5]), int(date[0:2]))
            except Exception:
                self.__get_view().error_message("Error 4: Incorrect format entered. Please try again.")
            else:
                break

        # day of the week
        day = date.weekday()
        match day:
            case 0:
                day = "Monday"
            case 1:
                day = "Tuesday"
            case 2:
                day = "Wednesday"
            case 3:
                day = "Thursday"
            case 4:
                day = "Friday"
            case 5:
                day = "Saturday"
            case 6:
                day = "Sunday"

        # suffix for the day of the month
        if day_number == 1 or day_number == 21 or day_number == 31:
            number_suffix = "st"
        elif day_number == 2 or day_number == 22:
            number_suffix = "nd"
        elif day_number == 3 or day_number == 23:
            number_suffix = "rd"
        else:
            number_suffix = "th"

        # month
        match month:
            case 1:
                month = "January"
            case 2:
                month = "February"
            case 3:
                month = "March"
            case 4:
                month = "April"
            case 5:
                month = "May"
            case 6:
                month = "June"
            case 7:
                month = "July"
            case 8:
                month = "August"
            case 9:
                month = "September"
            case 10:
                month = "October"
            case 11:
                month = "November"
            case 12:
                month = "December"

        # put it all together
        return f"{day} {day_number}{number_suffix} {month} {year}"



    def __number_of_groups(self, prefered_group_size: int) -> int | tuple[int, int]:
        '''
        Calculates the number of groups needed for the event
        :param preferred_group_size: integer representing the number preferred group size
        :return: tuple [number of groups, max group size]
        '''

        number_of_players = len(self.__get_model().get_players())
        # number_of_players = 37

        # Even number of people in each group
        if number_of_players % prefered_group_size == 0:
            return number_of_players // prefered_group_size, prefered_group_size

        if number_of_players // prefered_group_size == 1:
            return (number_of_players // prefered_group_size) + 1, prefered_group_size

        # Odd number of people in each group
        above_preferred_max = self.__get_view().change_group_size()
        if above_preferred_max.lower() == "y" or above_preferred_max.lower() == "yes":
            return number_of_players // prefered_group_size, prefered_group_size + 1
        else:
            return (number_of_players // prefered_group_size) + 1, prefered_group_size


    def __create_groups(self, number_of_groups: int, max_group_size: int) -> int:
        '''
        Creates the groups for the competition vai the sanke method
        :param number_of_groups: the number of groups
        :param max_group_size: the max size of a group
        :returns: The largest group size
        '''


        # Initializes the groups
        groups = []
        group_length = [0] * number_of_groups
        for i in range(number_of_groups):
            groups.append([])

        previous_group_number = 0
        self.__get_model().set_group_number(0)
        clash_moved_to = 1234567890

        players = self.__get_model().get_players()
        for player in players:
            # need to get it to skip the group where it placed the clashed person
            # check the previous group
            if clash_moved_to == self.__get_model().get_group_number():
                clash_moved_to = 123456789
                self.__change_group(number_of_groups)

            # if the current group is the same as the previous group then skip it

            # Check if there is a county clash
            clash = False
            for player_in_group in groups[self.__get_model().get_group_number()]:
                if player_in_group[2] == player[2]:
                    clash = True
                    # print(f"There are {group_length[self.group_number]} players")
                    # print(f"clash in group {self.group_number}")
                    break

            # Trys to change the group it is in when there is a clash
            if clash:
                original_group = self.__get_model().get_group_number()
                original_forward = self.__get_model().get_forward()
                original_end = self.__get_model().get_end()
                tries = 0
                while clash:
                    # move to the next group
                    self.__change_group(number_of_groups)
                    if self.__get_model().get_group_number() == original_group:
                        continue

                    # Makes sure that the next group isn't full
                    if len(groups[self.__get_model().get_group_number()]) != max_group_size:

                        # Checks for clash in new group
                        clash_new_group = False
                        for player_in_group in groups[self.__get_model().get_group_number()]:
                            if player_in_group[2] == player[2]:
                                clash_new_group = True
                                # print(f"clash in group {self.group_number}")

                        if not clash_new_group:
                            groups[self.__get_model().get_group_number()].append(player)
                            group_length[self.__get_model().get_group_number()] += 1
                            break
                        else:
                            tries += 1
                            if tries == number_of_groups - 1:
                                groups[original_group].append(player)
                                group_length[original_group] += 1
                                break
                    else:
                        tries += 1

                clash_moved_to = self.__get_model().get_group_number()
                self.__get_model().set_group_number(original_group)
                self.__get_model().set_forward(original_forward)
                self.__get_model().set_end(original_end)


            else:
                # Makes sure the group isnt full
                while (group_length[self.__get_model().get_group_number()] == max_group_size):
                    previous_group_number = self.__get_model().get_group_number()
                    self.__change_group(number_of_groups)

                # Adds the player to the group
                groups[self.__get_model().get_group_number()].append(player)
                group_length[self.__get_model().get_group_number()] += 1
                previous_group_number = self.__get_model().get_group_number()
                self.__change_group(number_of_groups)

        self.__get_model().set_groups(groups)
        return max(group_length)



    def __change_group(self, number_of_groups: int) -> None:
        '''
        Changes the group number in accordance with the snake method
        :param group_number: The current group number
        :param forward: The current direction the snake is going
        :param end: Whether the snake has done the end group twice or not
        :param number_of_groups: how many groups are in the competition
        :param clash: whether it is in the clash state or not
        :return: The next group number
        '''
        # change the group number
        if self.__get_model().get_forward():
            if self.__get_model().get_group_number() == (number_of_groups - 1) and self.__get_model().get_end() == 2:
                self.__get_model().set_forward(False)
                self.__get_model().set_group_number(self.__get_model().get_group_number() - 1)
                self.__get_model().set_end(1)
            elif self.__get_model().get_group_number() == (number_of_groups - 1) and self.__get_model().get_end() == 1:
                self.__get_model().set_end(2)
            else:
                self.__get_model().set_group_number(self.__get_model().get_group_number() + 1)
        else:
            if self.__get_model().get_group_number() == 0 and self.__get_model().get_end() == 2:
                self.__get_model().set_forward(True)
                self.__get_model().set_group_number(self.__get_model().get_group_number() + 1)
                self.__get_model().set_end(1)
            elif self.__get_model().get_group_number() == 0 and self.__get_model().get_end() == 1:
                self.__get_model().set_end(2)
            else:
                self.__get_model().set_group_number(self.__get_model().get_group_number() - 1)



    def __write_excel(self, group_file, sheet: xlsxwriter.workbook.Worksheet, number_of_groups: int,
                      largest_group_size: int, event: str) -> None:
        '''
        Writes the groups to the file
        :param groups_file:
        :param sheet:
        :param number_of_groups:
        :param largest_group_size:
        :param event:
        :return:
        '''
        # All the formats needed
        heading_format = group_file.add_format(
            {"bold": True, "border": 1, "top": 2, "align": "center", "valign": "vcenter"})
        date_event_format = group_file.add_format({"bold": True, "border": 1, "valign": "vcenter"})
        number_format = group_file.add_format({"bold": True, "border": 1, "align": "center", "valign": "vcenter"})
        licence_number_format = group_file.add_format({"border": 1, "align": "center", "valign": "vcenter"})
        player_name_format = group_file.add_format({"border": 1, "valign": "vcenter"})

        date_event_last_group_format = group_file.add_format({"bold": True, "border": 1, "bottom": 2, "valign": "vcenter"})
        number_last_group_format = group_file.add_format(
            {"bold": True, "border": 1, "bottom": 2, "align": "center", "valign": "vcenter"})
        licence_number_last_group_format = group_file.add_format(
            {"border": 1, "bottom": 2, "align": "center", "valign": "vcenter"})
        player_name_last_group_format = group_file.add_format({"border": 1, "bottom": 2, "valign": "vcenter"})
        player_name_last_group_name_format = group_file.add_format(
            {"border": 1, "bottom": 2, "right": 2, "valign": "vcenter"})
        player_name_last_name_format = group_file.add_format({"border": 1, "right": 2, "valign": "vcenter"})

        heading_end_format = group_file.add_format(
            {"bold": True, "border": 1, "top": 2, "right": 2, "align": "center", "valign": "vcenter"})

        # Creates the headings
        row = self.__get_model().get_rows()
        sheet.set_row(row, 19.5)
        sheet.write_string(row, 0, "Date", cell_format=heading_format)
        sheet.write_string(row, 1, "Event", cell_format=heading_format)
        sheet.write_string(row, 2, "Group", cell_format=heading_format)

        col = 3
        for i in range(largest_group_size):
            if self.__get_model().get_player_numbers():
                sheet.write_string(row, col, f"No{chr(i + 65)}", cell_format=heading_format)
                col += 1
            sheet.write_string(row, col, f"Cod{chr(i + 65)}", cell_format=heading_format)
            col += 1
            sheet.write_string(row, col, f"Player{chr(i + 65)}", cell_format=heading_format)
            col += 1
            sheet.write_string(row, col, f"c{chr(i + 65)}",
                               cell_format=heading_format if i + 1 != largest_group_size else heading_end_format)
            col += 1

        row += 1

        # writes the date
        sheet.write_string(row, 0, self.__get_model().get_long_date(), cell_format=date_event_format)

        # writes the Groups
        for i in range(number_of_groups):
            if i != 0:
                sheet.write_string(row, 0, "",
                                   cell_format=player_name_format if i + 1 != number_of_groups else player_name_last_group_format)
            sheet.write_string(row, 1, event,
                               cell_format=date_event_format if i + 1 != number_of_groups else date_event_last_group_format)
            sheet.write_number(row, 2, i + 1,
                               cell_format=number_format if i + 1 != number_of_groups else number_last_group_format)
            col = 3

            player_count = 0
            groups = self.__get_model().get_groups()
            for player in groups[i]:
                # writes the player number
                if self.__get_model().get_player_numbers():
                    sheet.write_number(row, col, 1,
                                       cell_format=number_format if i + 1 != number_of_groups else number_last_group_format)
                    col += 1

                # writes the licence number
                sheet.write_number(row, col, player[0],
                                   cell_format=licence_number_format if i + 1 != number_of_groups else licence_number_last_group_format)
                col += 1

                # writes player name
                sheet.write_string(row, col, player[1],
                                   cell_format=player_name_format if i + 1 != number_of_groups else player_name_last_group_format)
                col += 1
                player_count += 1

                # writes the county
                if player_count == largest_group_size:
                    format = player_name_last_name_format if i + 1 != number_of_groups else player_name_last_group_name_format
                else:
                    format = player_name_format if i + 1 != number_of_groups else player_name_last_group_format
                sheet.write_string(row, col, player[2], cell_format=format)
                col += 1

            # fills in the blanks if needed
            while player_count != largest_group_size:
                if self.__get_model().get_player_numbers():
                    sheet.write_string(row, col, "",
                                       cell_format=number_format if i + 1 != number_of_groups else number_last_group_format)
                    col += 1

                sheet.write_string(row, col, "",
                                   cell_format=licence_number_format if i + 1 != number_of_groups else licence_number_last_group_format)
                col += 1

                sheet.write_string(row, col, "",
                                   cell_format=player_name_format if i + 1 != number_of_groups else player_name_last_group_format)
                col += 1

                if player_count == largest_group_size - 1:
                    format = player_name_last_name_format if i + 1 != number_of_groups else player_name_last_group_name_format
                else:
                    format = player_name_format if i + 1 != number_of_groups else player_name_last_group_format
                sheet.write_string(row, col, "", cell_format=format)
                col += 1

                player_count += 1

            row += 1
        row += 1

        self.__get_model().set_rows(row)



    def __group_sheets(self, event: str) -> None:
        '''
        Creates the groups sheets
        :param event:
        :return:
        '''

        try:
            os.listdir("output files/group sheets")
        except FileNotFoundError:
            os.mkdir("output files/group sheets")

        # creates the event folder if not exists
        try:
            shutil.rmtree(f"output files/group sheets/{event}")
        except FileNotFoundError:
            os.mkdir(f"output files/group sheets/{event}")
        else:
            os.mkdir(f"output files/group sheets/{event}")

        # creates the sheets for all the groups in the event
        groups = self.__get_model().get_groups()
        for i, group in enumerate(groups):
            group_size = len(groups[i])

            # if self.player_numbers:
            #     pass
            # else:
            #     shutil.copy(f"assets/group_sheets/{group_size}.docx", f"output files/group sheets/{event}/{i + 1}.docx")

            # creates the keys for the dictionay
            keys = ["competition_name", "event_name", "event_date", "group_number"]
            for x in range(group_size):
                if self.__get_model().get_player_numbers():
                    keys.append(f"player_number{chr(x + 65)}")
                keys.append(f"cod{chr(x + 65)}")
                keys.append(f"Player{chr(x + 65)}")
                keys.append(f"c{chr(x + 65)}")

            # puts the data into a single list
            data = [self.__get_model().get_competition_name(), event, self.__get_model().get_short_date(), str(i + 1)]
            for player in group:
                if self.__get_model().get_player_numbers():
                    pass
                data.append(player[0])
                data.append(player[1])
                data.append(player[2])

            # creates the dictionary
            dictionary = {}
            for x, key in enumerate(keys):
                dictionary[key] = data[x]

            # edits the file accordingly
            document = DocxTemplate(f"resources/group_sheets/{group_size}.docx")

            document.render(dictionary)
            document.save(f"output files/group sheets/{event}/{i + 1}.docx")



    def __groups(self):
        # Loads the seeding File
        seeding_directory = "output files/seedings.xlsx"
        try:
            player_excel = load_workbook(seeding_directory)
        except FileNotFoundError:
            self.__get_view().error_message("Error code 3: Seeding list not Found")
            self.__menu_choice(1)

        # Inputs
        self.__get_model().set_competition_name(self.__get_view().competition_name())

        # while True:
        #     player_numbers = input("Are player numbers in use at this event? (y/n): ")
        #     if player_numbers.lower() == "y" or player_numbers.lower() == "yes":
        #         self.player_numbers = True
        #         break
        #     elif player_numbers.lower() == "n" or player_numbers.lower() == "no":
        #         break
        #     else:
        #         print("Sorry, I didn't understand")

        # grabs its Sheets and basic layout setup
        events = load_workbook(seeding_directory).sheetnames

        groups = xlsxwriter.Workbook("output files/Groups.xlsx")
        sheet = groups.add_worksheet("Groups")
        sheet.set_landscape()
        sheet.merge_range("A1:G1", f"{self.__get_model().get_competition_name()}",
                          cell_format=groups.add_format({'bold': True, "valign": "vcenter"}))
        sheet.set_row(0, 19.5)

        # Looks at all the sheets in the Excel file
        for idx, event in enumerate(events):
            self.__get_model().set_players([])
            # Date
            self.__get_model().set_long_date(self.__date(event))

            # Player info
            self.__get_player_list(player_excel, event)

            # Group size
            self.__get_view().message(f"There are {len(self.__get_model().get_players())} in this event")
            preferred_group_size = self.__get_view().group_size()

            # Calculate number of groups
            number_of_groups, max_group_size = self.__number_of_groups(preferred_group_size)

            # Create Groups
            largest_group_size = self.__create_groups(number_of_groups, max_group_size)

            # Writes groups to excel
            self.__write_excel(groups, sheet, number_of_groups, largest_group_size, event)
            self.__group_sheets(event)

        sheet.autofit()
        groups.close()
        player_excel.close()

        self.__menu_choice(1)



#================================================= Getters and Setters =================================================



    def __get_model(self):
        return self.model

    def __get_view(self):
        return self.view

    def __get_engine(self):
        return self.engine