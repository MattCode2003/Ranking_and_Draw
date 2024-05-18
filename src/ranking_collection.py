'''
Created on March 27, 2024
@author: Matthew Pryke
'''



#================================== Imports ===================================



import pandas as pd
from openpyxl.workbook import Workbook
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
import xlsxwriter
from xlsxwriter.workbook import Workbook as xlsxWorkbook
import os
import time
from utils.text_colour import bcolors
import main


class RankingCollection:
    def __init__(self) -> None:
        self.player_directory = "input files/player_list.xlsx"

        ranking_directory = ""
        input_files_directory = str(os.path.join(os.getcwd(), "input files"))

        for filename in os.listdir(os.path.join(os.getcwd(), "input files")):
            if "Alpha".lower() in filename.lower():
                ranking_directory = f"{input_files_directory}\\{filename}"
        if ranking_directory == "":
            print(f"{bcolors.WARNING}Error code 1: Alpha list not Found{bcolors.ENDC}")
            time.sleep(3)
            exit(1)

        # Turns Ranking List Excel File into Database stored in RAM
        self.engine = create_engine('sqlite:///', echo=False)
        df = pd.read_excel(ranking_directory)
        df.to_sql('ranking_list', con=self.engine, if_exists='replace', index=False)


        # Turns the County codes excel file into database stored in RAM
        try:
            df = pd.read_excel("resources/County Codes.xlsx")
        except FileNotFoundError:
            print("Error code 2: County_Codes file not found")
            time.sleep(3)
            exit(2)
        finally:
            df.to_sql('county_codes', con=self.engine, if_exists='replace', index=False)

        self.players = []


    def __get_player_list(self, player_excel: Workbook, event: str) -> None:
        ws = player_excel[event]

        # Skip age group if it is not taking place
        if ws["A2"].value is None:
            return

        # In every Sheet it takes all the names and licence numbers in each sheet and puts it in a list
        counter = 2
        while True:
            person = []
            if ws["A" + str(counter)].value is not None:
                person.append(ws["A" + str(counter)].value)
                person.append(ws["B" + str(counter)].value)
                person.append(ws["C" + str(counter)].value)
                self.players.append(person)
                counter += 1
            else:
                break



    def __event(self, event: str) -> str:
        # Works out what Age Category the Players are in
        match event:
            case "u11b":
                current_event = "Under 11 Boys"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u11g":
                current_event = "Under 11 Girls"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u13b":
                current_event = "Under 13 Boys"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u13g":
                current_event = "Under 13 Girls"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "cadet boys":
                current_event = "Under 15 Boys"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "cadet girls":
                current_event = "Under 15 Girls"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "jnr boys":
                current_event = "Under 19 Men"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "jnr girls":
                current_event = "Under 19 Women"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u21m":
                current_event = "Under 21 Men"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "u21w":
                current_event = "Under 21 Women"
                for player in self.players:
                    player.append(self.__get_points(player[0], current_event, ""))

            case "mens singles":
                current_event = "Mens Singles"
                for player in self.players:
                    player.append(self.__get_points(player[0], "Men", "Senior"))

            case "womens singles":
                current_event = "Womens Singles"
                for player in self.players:
                    player.append(self.__get_points(player[0], "Women", "Senior"))

            case "mens vets":
                current_event = "Mens Vets Singles"
                for player in self.players:
                    player.append(self.__get_points(player[0], "Men", "Veteran"))

            case "womens vets":
                current_event = "Womens Vets Singles"
                for player in self.players:
                    player.append(self.__get_points(player[0], "Women", "Veteran"))

        return current_event



    def __get_points(self, licence_number: str, event: str, senior_category: str) -> int:
        with self.engine.connect() as conn:
            if senior_category != "":
                query = f"SELECT Points FROM ranking_list WHERE Membership_no = '{licence_number}' AND Sub_Category = '{event}' AND Category = '{senior_category}';"
            else:
                query = f"SELECT Points FROM ranking_list WHERE Membership_no = '{licence_number}' AND Sub_Category = '{event}';"
            result = conn.execute(text(query))
            points = result.fetchall()

            if points == []:
                just_points = "0"
            else:
                just_points = ""
                rb = list(str(points))
                for character in rb:
                    if character.isnumeric():
                        just_points += character

        return int(just_points)



    def __county_codes(self) -> None:
        # Converts the counties into the county codes

        for player in self.players:
            with self.engine.connect() as conn:
                query = f"SELECT Code FROM county_codes WHERE County = '{player[2]}';"
                result = conn.execute(text(query))
                county_code = result.fetchall()

            # Removes the extra crap that is in the county code string
            just_county_code = ""
            scounty = list(str(county_code))
            for char in scounty:
                if char.isalpha():
                    just_county_code += char

            player[2] = just_county_code



    def __write_to_file(self, seedings: xlsxWorkbook, current_event: str) -> None:
        sheet = seedings.add_worksheet(current_event)

        sheet.write_string(0, 0, "Licence Number")
        sheet.write_string(0, 1, "Name")
        sheet.write_string(0, 2, "County")
        sheet.write_string(0, 3, "Points")
        for i, player in enumerate(self.players):
            for x, data in enumerate(player):
                if x == 0 or x == 3:
                    sheet.write_number(i + 1, x, data)
                else:
                    sheet.write_string(i + 1, x, data)

        # Auto changes the size of the cells to fit perfectly
        sheet.autofit()


    def main(self) -> None:
        # Loads the Player List File and grabs its Sheets
        player_excel = load_workbook(self.player_directory)
        events = load_workbook(self.player_directory).sheetnames

        # Creates the excel file for the seedings
        seedings = xlsxwriter.Workbook("output files/seedings.xlsx")

        # Looks at all the sheets in the Excel file
        for idx, event in enumerate(events):

            # Skip age group if it is not taking place
            ws = player_excel[event]
            if ws["A2"].value is None:
                continue

            self.players = []
            self.__get_player_list(player_excel, event)
            current_event = self.__event(event)

            # Sorts the players in order of most points
            self.players.sort(key=lambda x: x[3], reverse=True)
            self.__county_codes()
            self.__write_to_file(seedings, current_event)

        # if there is no output folder then one is created
        path = "output files"
        if not os.path.exists(path):
            os.makedirs(path)
        seedings.close()
        player_excel.close()

        main.Main().menu()



if __name__ == "__main__":
    RankingCollection.main()